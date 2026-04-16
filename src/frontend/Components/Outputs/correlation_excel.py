"""
Applies recode settings to a SAV dataframe and generates a correlation table as a downloadable Excel file.
Meant to be called after render_sav_processor() in main.py.
"""
import io
import pandas as pd
import streamlit as st


def _convert_value(value, settings: dict):
    """
    Apply recode logic to a single value based on its settings.
    Handles NaN/missing values via sysmis_becomes if set.
    Returns recoded integer (1 or 2), or None if value doesn't match or becomes is None.
    """
    variable_type = settings.get('variable_type', 'categorical')
    sysmis_becomes = settings.get('sysmis_becomes')

    if pd.isna(value):
        return sysmis_becomes  # None if not set, or 1/2 if user mapped it

    if variable_type == 'unknown':
        return None

    if variable_type == 'continuous':
        op1 = settings['range1_operator']
        val1 = settings['range1_value']
        becomes1 = settings['range1_becomes']

        if becomes1 is not None:
            if   op1 == '<'  and value <  val1: return becomes1
            elif op1 == '<=' and value <= val1: return becomes1
            elif op1 == '='  and value == val1: return becomes1
            elif op1 == '>=' and value >= val1: return becomes1
            elif op1 == '>'  and value >  val1: return becomes1

        op2 = settings['range2_operator']
        val2 = settings['range2_value']
        becomes2 = settings['range2_becomes']

        if becomes2 is not None:
            if   op2 == '<'  and value <  val2: return becomes2
            elif op2 == '<=' and value <= val2: return becomes2
            elif op2 == '='  and value == val2: return becomes2
            elif op2 == '>=' and value >= val2: return becomes2
            elif op2 == '>'  and value >  val2: return becomes2

        return None

    else:
        r1_start = settings.get('range1_start')
        r1_end = settings.get('range1_end')
        r2_start = settings.get('range2_start')
        r2_end = settings.get('range2_end')
        r1_becomes = settings.get('range1_becomes')
        r2_becomes = settings.get('range2_becomes')

        if None not in (r1_start, r1_end) and r1_becomes is not None:
            if r1_start <= value <= r1_end:
                return r1_becomes

        if None not in (r2_start, r2_end) and r2_becomes is not None:
            if r2_start <= value <= r2_end:
                return r2_becomes

        return None


def apply_recodes(df: pd.DataFrame, recode_settings: dict) -> tuple[pd.DataFrame, list[str]]:
    """
    Apply all recode settings to the dataframe, adding Recode: columns.
    Returns tuple of (recoded dataframe, list of skipped question labels).
    """
    df = df.copy()
    skipped = []

    for label, settings in recode_settings.items():
        column = settings.get('column') or settings.get('matched_column')

        if not column:
            skipped.append(f"{label} (no SAV column matched)")
            continue
        if column not in df.columns:
            skipped.append(f"{label} (column '{column}' not found in SAV)")
            continue
        if settings.get('variable_type') == 'unknown':
            skipped.append(f"{label} (variable type could not be determined)")
            continue
        # Skip only if all three are None
        if (settings.get('range1_becomes') is None and
                settings.get('range2_becomes') is None and
                settings.get('sysmis_becomes') is None):
            skipped.append(f"{label} (all ranges set to None)")
            continue

        df[f"Recode: {label}"] = df[column].apply(lambda x: _convert_value(x, settings))

    return df, skipped


def build_correlation_table(df: pd.DataFrame, recode_settings: dict) -> pd.DataFrame:
    """Build a correlation matrix from all recoded label columns, deduplicating first."""
    label_cols = list(dict.fromkeys([
        f"Recode: {label}" for label in recode_settings.keys()
        if f"Recode: {label}" in df.columns
    ]))

    if not label_cols:
        return pd.DataFrame()

    return df[label_cols].corr().abs()


def _write_styled_excel(corr_matrix: pd.DataFrame) -> io.BytesIO:
    """Write correlation matrix to a styled Excel buffer matching SPSS style."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    BLUE_FONT   = Font(color="1F3864", bold=False)
    BLUE_BOLD   = Font(color="1F3864", bold=True)
    GRAY_FILL   = PatternFill("solid", fgColor="D9D9D9")
    WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
    HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
    THIN_BORDER = Border(
        bottom=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF")
    )
    WRAP = Alignment(wrap_text=True, vertical="center", horizontal="center")
    WRAP_LEFT = Alignment(wrap_text=True, vertical="center", horizontal="left")

    wb = Workbook()
    ws = wb.active
    ws.title = "Correlation Table"

    labels = list(corr_matrix.columns)
    n = len(labels)

    ws.cell(row=1, column=1, value="")
    for col_idx, label in enumerate(labels, start=2):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = BLUE_BOLD
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        cell.border = THIN_BORDER

    for row_idx, label in enumerate(labels, start=2):
        is_gray = (row_idx % 2 == 0)
        row_fill = GRAY_FILL if is_gray else WHITE_FILL

        idx_cell = ws.cell(row=row_idx, column=1, value=label)
        idx_cell.font = BLUE_BOLD
        idx_cell.fill = HEADER_FILL
        idx_cell.alignment = WRAP_LEFT
        idx_cell.border = THIN_BORDER

        for col_idx, col_label in enumerate(labels, start=2):
            try:
                val = corr_matrix.loc[label, col_label]
            except Exception:
                val = None
            cell = ws.cell(row=row_idx, column=col_idx, value=round(val, 3) if val is not None and not pd.isna(val) else "")
            cell.font = BLUE_FONT
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 35
    for col_idx in range(2, n + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    ws.row_dimensions[1].height = 120
    for row_idx in range(2, n + 2):
        ws.row_dimensions[row_idx].height = 120

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def render_correlation_exporter():
    """
    Streamlit component that applies recodes, builds correlation table,
    and renders a download button for the Excel file.
    """
    sav_data = st.session_state.get('sav_data')
    recode_settings = st.session_state.get('recode_settings')

    if not sav_data or not recode_settings:
        return

    df = sav_data['df']

    try:
        recoded_df, skipped = apply_recodes(df, recode_settings)
        corr_matrix = build_correlation_table(recoded_df, recode_settings)

        if corr_matrix.empty:
            st.warning("⚠️ No recoded columns found to correlate.")
            return

        buffer = _write_styled_excel(corr_matrix)

        st.subheader("📈 Correlation Table")
        st.download_button(
            label="📥 Download Correlation Table (.xlsx)",
            data=buffer,
            file_name="correlation_table.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )

        if skipped:
            with st.expander(f"⚠️ {len(skipped)} question(s) excluded from correlation table"):
                for reason in skipped:
                    st.caption(f"• {reason}")

    except Exception as e:
        st.error(f"❌ Error generating correlation table: {str(e)}")
        st.exception(e)